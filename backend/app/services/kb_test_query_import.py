"""Parse bulk test-query imports (CSV / XLSX) into normalized row dicts.

Evaluators keep complete validation test sets in spreadsheets; this module
turns an uploaded file into rows the import endpoint can upsert as
``KBTestQuery`` records. Pure parsing — no DB access — so it unit-tests
without Beanie.
"""

import csv
import datetime
import io
import logging

logger = logging.getLogger(__name__)

MAX_IMPORT_ROWS = 500

# Canonical field -> accepted header spellings (normalized: lowercased,
# underscores/extra whitespace collapsed to single spaces).
_HEADER_ALIASES: dict[str, set[str]] = {
    "query": {"question", "query", "prompt", "test question", "test query"},
    "expected_answer": {
        "expected answer", "answer", "expected response", "canonical answer",
    },
    "category": {
        "category", "type", "question type", "question category",
        "category/type", "question category/type",
    },
    "expected_source_labels": {
        "source", "sources", "section", "source or section", "source/section",
        "expected sources", "source labels", "expected source labels",
    },
    "notes": {"notes", "note", "comments", "comment"},
    "external_id": {
        "id", "question id", "external id", "stable id", "stable question id",
        "qid",
    },
    "expected_answer_contains": {"expected answer contains", "answer contains"},
}


class TestQueryImportError(ValueError):
    """File-level problem the user can act on (bad format, missing columns)."""



def _split_source_labels(raw: str) -> list[str]:
    """Split a source-label cell into individual labels.

    Commas are the obvious separator, but real source names contain them —
    "Subpart D-iii — Monitoring, Reporting, Remedies & Closeout" is one
    source, not three. Splitting it blindly invents labels that match no
    source in the KB, and every invented label also inflates the denominator
    of the retrieval-precision score, so the KB gets marked down for a
    spreadsheet formatting artifact rather than a retrieval failure.

    Two escape hatches let an author be unambiguous:

      * Quotes protect their contents — ``"Monitoring, Reporting"`` is one
        label. Straight and curly quotes both work, since spreadsheets like
        to autocorrect them.
      * Semicolons, when the cell contains any, become the *only* separator,
        leaving commas inside each label intact.

    A cell with neither still splits on commas, which keeps every previously
    valid file importing exactly as before.
    """
    if not raw or not raw.strip():
        return []

    # Semicolons are the explicit separator: once present, commas are content.
    separators = {";"} if ";" in raw else {";", ","}
    closing = {'"': '"', "\u201c": "\u201d", "\u201d": "\u201d"}

    labels: list[str] = []
    buf: list[str] = []
    expected_close: str | None = None

    for ch in raw:
        if expected_close is not None:
            if ch == expected_close:
                expected_close = None
            else:
                buf.append(ch)
        elif ch in closing:
            expected_close = closing[ch]
        elif ch in separators:
            labels.append("".join(buf))
            buf = []
        else:
            buf.append(ch)
    labels.append("".join(buf))

    return [label.strip() for label in labels if label.strip()]


def _normalize_header(value: str) -> str:
    return " ".join(value.replace("_", " ").strip().lower().split())


def _cell_to_str(value) -> str:
    """Coerce a spreadsheet cell to clean text.

    Excel stores bare numbers as floats, so an ID column of ``1, 2, 3`` reads
    back as ``1.0, 2.0, 3.0`` — strip the spurious decimal.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.isoformat()
    return str(value).strip()


def _raw_rows_from_csv(data: bytes) -> list[list[str]]:
    # utf-8-sig eats the BOM Excel prepends to exported CSVs; Windows-Excel
    # exports may instead be cp1252, so fall back rather than reject the file.
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("cp1252", errors="replace")
    return [row for row in csv.reader(io.StringIO(text))]


def _raw_rows_from_xlsx(data: bytes) -> list[list[str]]:
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        logger.warning("Test-query import: unreadable xlsx: %s", e)
        raise TestQueryImportError(
            "Couldn't read the Excel file. Re-save it as .xlsx (or export as CSV) and try again."
        )
    try:
        ws = wb.active
        if ws is None:
            return []
        return [[_cell_to_str(c) for c in row] for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def parse_test_query_import(filename: str, data: bytes) -> tuple[list[dict], list[dict]]:
    """Parse an uploaded CSV/XLSX into (rows, row_errors).

    Each row dict carries the ``KBTestQuery`` user fields: ``query``,
    ``expected_answer``, ``expected_answer_contains``, ``category``,
    ``expected_source_labels`` (list), ``notes``, ``external_id`` — absent
    columns yield ``None``/``[]``. ``row_errors`` entries are
    ``{"row": <1-based spreadsheet row>, "error": str}`` for rows that were
    skipped but shouldn't fail the whole import.

    Raises ``TestQueryImportError`` for file-level problems.
    """
    lowered = (filename or "").lower()
    if lowered.endswith(".csv"):
        raw = _raw_rows_from_csv(data)
    elif lowered.endswith(".xlsx"):
        raw = _raw_rows_from_xlsx(data)
    elif lowered.endswith(".xls"):
        raise TestQueryImportError(
            "Legacy .xls files aren't supported. Re-save the spreadsheet as .xlsx or .csv."
        )
    else:
        raise TestQueryImportError(
            "Unsupported file type. Upload a .csv or .xlsx file."
        )

    # Drop fully-empty leading rows so a title row of blanks doesn't become
    # the header. header_offset keeps reported row numbers matching the sheet.
    header_offset = 0
    while raw and not any(c.strip() for c in raw[0] if isinstance(c, str)):
        raw.pop(0)
        header_offset += 1
    if not raw:
        raise TestQueryImportError("The file is empty.")

    header = raw[0]
    columns: dict[int, str] = {}
    for idx, cell in enumerate(header):
        normalized = _normalize_header(str(cell or ""))
        for field, aliases in _HEADER_ALIASES.items():
            if normalized in aliases and field not in columns.values():
                columns[idx] = field
                break
    if "query" not in columns.values():
        raise TestQueryImportError(
            'No question column found. The header row must include a "Question" '
            '(or "Query") column. Download the template for the expected format.'
        )

    data_rows = raw[1:]
    if len(data_rows) > MAX_IMPORT_ROWS:
        raise TestQueryImportError(
            f"The file has {len(data_rows)} rows; the limit is {MAX_IMPORT_ROWS} "
            "per import. Split it into smaller files."
        )

    rows: list[dict] = []
    errors: list[dict] = []
    for i, raw_row in enumerate(data_rows):
        # 1-based sheet row: skipped blanks + header + this row.
        sheet_row = header_offset + 1 + i + 1
        values = {field: "" for field in _HEADER_ALIASES}
        for idx, field in columns.items():
            if idx < len(raw_row):
                values[field] = _cell_to_str(raw_row[idx])
        if not any(v for v in values.values()):
            continue  # blank spacer row
        if not values["query"]:
            errors.append({"row": sheet_row, "error": "Missing question"})
            continue
        labels = _split_source_labels(values["expected_source_labels"])
        rows.append({
            "row": sheet_row,
            "query": values["query"],
            "expected_answer": values["expected_answer"] or None,
            "expected_answer_contains": values["expected_answer_contains"] or None,
            "category": values["category"].lower() or None,
            "expected_source_labels": labels,
            "notes": values["notes"] or None,
            "external_id": values["external_id"] or None,
        })
    return rows, errors
